# -*- coding: utf-8 -*-
"""
excel文件读取与写入


load excel参数
    read_only：是否为只读模式，对于超大型文件，要提升效率有帮助
    keep_vba ：是否保留 vba 代码，即打开 Excel 文件时，开启并保留宏
    guess_types：是否做在读取单元格数据类型时，做类型判断
    data_only：是否将公式转换为结果，即包含公式的单元格，是否显示最近的计算结果
    keep_links：是否保留外部链接



单元格格式styles:
    NumberFormat 数字
    Alignment 对齐
    Font 字体
    Border 边框
    PatternFill 填充
    Protection 保护
"""
import os

import numpy as np
import openpyxl

def getSheetNames(file: str) -> list:
    """
    获取xlsx文件的sheet所有表格名

    :param file: xlsx文件路径
    :return: 表格名list
    """
    if not os.path.exists(file):
        raise FileNotFoundError(f'{file}文件不存在')
    try:
        workbook = openpyxl.load_workbook(file)
        sheetNames = list(workbook.sheetnames)
        print(f"读取{file}成功")
        workbook.close()
        return sheetNames
    except Exception as e:
        raise Exception(f"读取{file}失败：{e}")


def readSheet(file: str, sheetName: str=None,onlyStr:bool=False) -> [[str]]:
    """
    读取xlsx文件的sheet表格，返回二维列表，元素都转化为str类型

    :param file: xlsx文件路径
    :param sheetName: 文件内的表格名，没有则读取第一个
    :param onlyStr: 是否将结果转化为str类型
    :return: 表格数据list[list[]]
    """
    if not os.path.exists(file):
        raise FileNotFoundError(f'{file}文件不存在')
    try:
        workbook = openpyxl.load_workbook(file, read_only=False, data_only=True)
        sheetNames = list(workbook.sheetnames)
        if sheetName not in sheetNames:
            print(f"{file}文件中没有{sheetName}表格，将读取第一个表格: {sheetNames[0]}")
            sheetName = sheetNames[0]
        sheet = workbook[sheetName]
        data = []
        # 遍历工作表的每一行
        for row in sheet.iter_rows(values_only=True):
            row_data = []
            for cell_value in row:
                # 将单元格值转换为字符串，若为 None 则转换为空字符串
                if onlyStr:
                    cell_value = str(cell_value) if cell_value is not None else ""
                row_data.append(cell_value)
            data.append(row_data)
        workbook.close()
        print(f"读取{file}的{sheetName}成功")
        workbook.close()
        return data
    except Exception as e:
        raise Exception(f"读取{file}失败：{e}")


def appendSheet(file: str, sheet: str, data: list[list]) -> bool:
    """
    将data写入xlsx文件的sheet表格
    如果不存在file，则会创建，有则追加sheet

    :param file: 文件路径
    :param sheet: 表格名称
    :param data: 二维列表数据
    :return: 成功与否
    """
    if not os.path.exists(file):
        # 新建表格
        print(f"文件不存在创建新的xlsx文件")
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet(sheet)
    else:
        workbook = openpyxl.load_workbook(file, read_only=False, data_only=True)
        sheets = list(workbook.sheetnames)
        if sheet not in sheets:
            sheet = workbook.create_sheet(sheet)# 新建表格
        else:
            raise ValueError(f"{file}文件中已经存在{sheet}表格")
    try:
        for i in data:
            sheet.append(i)
        workbook.save(file)
        print(f"写入'{file}'文件的{sheet}'表格'成功")
        workbook.close()
        return True
    except Exception as e:
        raise Exception(f"写入{file}失败：{e}")


if __name__ == '__main__':
    # print(getSheetNames('./file/excel1.xlsx'))
    # data = readSheet('./file/excel1.xlsx', 'Sheet1')
    # print(data)
    # writeSheet('./file/excel2.xlsx', 'mySheet1', [['1', '2', '3'], ['4', '5', '6']])
    pass