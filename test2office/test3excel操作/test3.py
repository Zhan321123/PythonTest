"""
转自：
https://zhuanlan.zhihu.com/p/351998173
"""
from openpyxl import Workbook

wb = Workbook()

ws = wb.active

# 直接索引操作表格的具体内容
ws['A1'] = 100
# append会在加在新的一行
ws.append([1, 2, 3, 4])
ws['B2'] = 'zhan'

wb.save('../file/excel3.xlsx')
