"""
操作单元格cell

索引操作
    sheet['A1'] = value
    ws.cell(row=int,column=int,value=value)
    sheet['A'] = values on A column
    sheet['1'] = values on 1 row
    sheet['A:C'] = values from A to C column
    sheet['A1:C3'] = values from A1 to C3 rows
    ws.iter_rows()获取整行数据
    ws.iter_cols()获取整列数据
    cell.value
    sheet.merge_cells('cells')，合并单元格
    sheet.unmerge_cells('cells')，解除合并
"""
from objprint import objprint
from openpyxl import load_workbook

path = '../file/excel1.xlsx'
wb = load_workbook(path,data_only=True)
ws = wb['Sheet1']

ws['A10'] = 'PE'
ws['B10'] = '100'
d = ws.cell(row=4,column=4,value=0)

print(ws['1'])
for cell in ws['A']:
    print(cell.value)
print('*'*10)
for row in ws['A1:B2']:
    for cell in row:
        print(cell.value)

for row in ws.iter_rows():
    for cell in row:
        print(cell.value)

columns = (a for a in ws.iter_cols())
print(columns)
# wb.save(path)