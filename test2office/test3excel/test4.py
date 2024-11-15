"""
获取表格的表名sheetnames

    一、load excel参数
read_only：是否为只读模式，对于超大型文件，要提升效率有帮助
keep_vba ：是否保留 vba 代码，即打开 Excel 文件时，开启并保留宏
guess_types：是否做在读取单元格数据类型时，做类型判断
data_only：是否将公式转换为结果，即包含公式的单元格，是否显示最近的计算结果
keep_links：是否保留外部链接
"""
from objprint import objprint
from openpyxl import load_workbook

wb = load_workbook('../file/month.xlsx',read_only=True)
print(wb.sheetnames)
objprint(wb)