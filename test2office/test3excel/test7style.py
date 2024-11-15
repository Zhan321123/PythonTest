"""
单元格格式styles

    styles:
NumberFormat 数字
Alignment 对齐
Font 字体
Border 边框
PatternFill 填充
Protection 保护
"""
from openpyxl import Workbook
from openpyxl.styles import *

wb = Workbook()
ws = wb.active
ws['A1'] = '宋体'
ws['A1'].font = Font(name='宋体',size=12,bold=True,color='FF0000')
ws['A2'] ='右对齐'
ws['A2'].alignment = Alignment(horizontal='right')
ws.cell(row=3, column=3, value='填充渐变色').fill = PatternFill(fill_type='solid', start_color='FF0000')
ws.cell(row=4, column=4, value='设置边线').border = Border(left=Side(border_style='thin', color='FF0000'), right= Side(border_style='thin', color='FF0000'))
ws.cell(row=5, column=5, value='受保护的').protection = Protection(locked=True, hidden=True)
ws.cell(row=6, column=6, value=0.54).number_format =numbers.FORMAT_PERCENTAGE

wb.save('../file/excel3.xlsx')
