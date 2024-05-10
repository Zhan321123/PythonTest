"""
操作sheet

Workboob.worksheets获取所有sheet对象
create_sheet('name',location=int)添加新sheet
sheet.title = str:name
Workbook.sheet_properties.tabColor = str:color，修改sheet颜色
依据name获取sheet:
    Workbook.get_sheet_by_name('sheet_name')
    sheet = Workbook['sheet']

Workbook.copy_worksheet(sheet)，复制sheet
Workbook.remove(name=str)，删除sheet

"""
from objprint import objprint
from openpyxl import load_workbook

wb = load_workbook('../file/excel1.xlsx')

ws = wb.active
objprint(ws)
# ws = wb['name']
ws2 = wb.create_sheet('sheet2')
ws2.title = u'新表单'
ws3 = wb.create_sheet('sheet0', 0)
ws3.title = 'mysheet0'

ws.sheet_properties.tabColor = '1072BA'

ws4 = wb.get_sheet_by_name('mysheet0')
print(ws4)

ws4_copy = wb.copy_worksheet(ws4)
print(wb.sheetnames)

wb.remove(ws2)
wb.remove(ws4)
wb.remove(ws3)


wb.save('../file/excel1.xlsx')