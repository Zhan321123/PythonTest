import os

import mysql.connector
from mysql.connector import Error
import numpy as np
import openpyxl


def readSheet(file: str, sheetName: str = None, onlyStr: bool = False) -> [[str]]:
  if not os.path.exists(file):
    raise FileNotFoundError(f'{file}文件不存在')
  try:
    workbook = openpyxl.load_workbook(file, read_only=False, data_only=True)
    sheetNames = list(workbook.sheetnames)
    if sheetName not in sheetNames:
      print(f"{file}文件中没有{sheetName}表格，将读取第一个表格: {sheetNames[0]}")
      sheetName = sheetNames[0]
    sheet = workbook[sheetName]
    data = []
    for row in sheet.iter_rows(values_only=True):
      row_data = []
      for cell_value in row:
        if onlyStr:
          cell_value = str(cell_value) if cell_value is not None else ""
        row_data.append(cell_value)
      data.append(row_data)
    workbook.close()
    print(f"读取{file}的{sheetName}成功")
    workbook.close()
    return data
  except Exception as e:
    raise Exception(f"读取{file}失败：{e}")

if __name__ == '__main__':
  data = np.array(readSheet(r"C:\Users\刘高瞻\Desktop\市政业绩.xlsx")).flatten()
  # 查看data是否有重复数据
  length,setLength = len(data),len(set(data))
  print(f"数据长度：{length}，去重后长度：{setLength}") # 数据长度：11913，去重后长度：7558
  if length != setLength:
    print("数据有重复数据")
    data = np.unique(data)
  # data = data[:10]
  print(max(len(i) for i in data)) #126
  try:
    # 创建数据库连接
    connection = mysql.connector.connect(
      host='127.0.0.1',
      database='test',
      user='root',
      password='987654321asd'
    )

    if connection.is_connected():
      print("数据库连接成功！")
      cursor = connection.cursor()

      cursor.execute("show databases")
      databases = cursor.fetchall()
      print("已存在的database：", databases)

      cursor.execute("show tables")
      tables = cursor.fetchall()
      print("已存在的表：", tables)
      # cursor.execute("drop table companyNames")
      cursor.execute("show tables")
      tables = cursor.fetchall()
      print("已存在的表：", tables)


      createTable = """
      create table companyNames(
        id int auto_increment primary key,
        name varchar(255) unique not null,
        index name_index (name)
    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
      """
      cursor.execute(createTable)
      print("创建表成功！")

      insertSql = "insert into companyNames(name) values(%s)"
      dataToInsert = [(i,) for i in data]
      cursor.executemany(insertSql, dataToInsert)

      # 查看表
      cursor.execute("select * from companyNames")
      print("表内容：", cursor.fetchall())

      connection.commit()
  except Error as e:
    print("数据库操作错误：", e)
  finally:
    # 关闭数据库连接
    if connection.is_connected():
      cursor.close()
      connection.close()
      print("数据库连接已关闭。")
