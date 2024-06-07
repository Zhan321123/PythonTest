# -*- coding: utf-8 -*-
import os
from typing import Sequence
import openpyxl
import pandas
import pandas as pd
from pandas import DataFrame


class ExcelReader:
    """
    Use pandas to build an Excel Reader
    """
    file: str
    workbook: dict[str, DataFrame]

    def __init__(self, file: str, header=None):
        self._excelChange(file, header)

    def _excelChange(self, file: str, header=None):
        if file.endswith('.xlsx'):
            self.file = file
            self.workbook = pandas.read_excel(self.file, sheet_name=None, header=header)
            self._sheets = self.workbook.keys()
        else:
            print('File is probably not an excel file')
            self.file = file + '.xlsx'

    def setFile(self, file: str):
        """
        Modify the Excel file path
        """
        self._excelChange(file)

    def getSheetNames(self):
        """
        Get this workbook's all sheets' names
        Return list of sheets' names
        """
        return list(self.workbook.keys())

    def getSheet(self, sheet: str):
        """
        Get a sheet's all cells
        """
        if sheet not in self._sheets:
            return None
        else:
            return self.workbook[sheet].to_numpy().tolist()

    def getCell(self, sheet: str, row: int, column: int):
        """
        Get sheet's the cell value
        Return the sheet's row and column's value
        """
        if sheet not in self._sheets:
            return None
        else:
            return self.workbook[sheet].iloc[row - 1, column - 1]

    def getRowData(self, sheet: str, row: int):
        """
        Get sheet's the row data
        The first row is sheet head
        So start from the second row
        """
        if sheet not in self._sheets:
            return None
        else:
            return self.workbook[sheet].iloc[row - 1].to_list()

    def getColumnData(self, sheet: str, column: int):
        """
        Get sheet's the column data
        """
        if sheet not in self._sheets:
            return None
        else:
            return self.workbook[sheet].iloc[:, column - 1].to_list()


class SingleSheetExcelReader(ExcelReader):
    """
    this excel only has one sheet
    if not,will only read the first sheet
    """

    def __init__(self, file: str, header=0):
        super().__init__(file, header)
        self.sheet = self.getSheetNames()[0]

    def getSheet(self):
        return super().getSheet(self.sheet)

    def getCell(self, row: int, column: int):
        return super().getCell(self.sheet, row, column)

    def getRowData(self, row: int):
        return super().getRowData(self.sheet, row)

    def getColumnData(self, column: int):
        return super().getColumnData(self.sheet, column)


class ExcelWriter:
    """
    Use pandas and openpyxl to build an Excel Writer
    """
    file: str
    workbook: openpyxl.Workbook

    def __init__(self, file: str):
        if file.endswith('.xlsx'):
            self.file = file
        else:
            self.file = file + '.xlsx'
        if os.path.exists(file):
            self.workbook = openpyxl.load_workbook(self.file, read_only=False, data_only=True)
        else:
            self.workbook = openpyxl.Workbook()
            self.workbook.save(self.file)

    def resetWorkbook(self, workbook: dict[str, DataFrame]):
        pass

    def appendSheet(self, sheet: str, data: DataFrame) -> bool:
        """
        append a sheet
        return dose the workbook has this sheet
        """
        if not self._sheetExists(sheet):
            df = pd.DataFrame(data)
            self.workbook.create_sheet(sheet)
            s = self.workbook.worksheets[sheet]
            for row in df.values:
                s.append(row.tolist())
            return True
        else:
            print('Sheet already exists')
            return False

    def modifyCellValue(self, sheet: str, row: int, column: int, value):
        """
        modify the workbook's sheet's row and colum's value
        """
        if not self._sheetExists(sheet):
            return
        self.workbook[sheet].cell(row=row, column=column).value = value

    def _sheetExists(self, sheet: str) -> bool:
        """
        this workbook has this sheet
        """
        return sheet in self.workbook.sheetnames

    def modifyRow(self, sheet: str, row: int, dataList: Sequence):
        """
        modify sheet's row's values
        """
        if not self._sheetExists(sheet):
            return
        for i in range(len(dataList)):
            self.workbook[sheet].cell(row=row, column=i + 1).value = dataList[i]

    def modifyColumn(self, sheet: str, column: int, dataList: Sequence):
        """
        modify sheet's column's values
        """
        if not self._sheetExists(sheet):
            return
        for i in range(len(dataList)):
            self.workbook[sheet].cell(row=i + 1, column=column).value = dataList[i]

    def save(self):
        """
        save workbook
        """
        self.workbook.save(self.file)

    def close(self):
        """
        close
        """
        self.workbook.close()


class CsvWriter:
    def __init__(self, data: list[list]):
        self.data = data

    def write(self, file: str = 'output.csv'):
        pandas.DataFrame(self.data).to_csv(file, header=False, index=False)


class CsvReader:
    def __init__(self, file: str,header=None):
        self.file = file
        self.csv = pandas.read_csv(file, header=header, encoding='gbk', low_memory=False)
        self.rawData = self.csv.to_numpy().tolist()

    def getData(self):
        return self.rawData

    def getColumn(self, column: int):
        return self.csv.iloc[:, column - 1].to_list()

    def getColumns(self, columns: list[int]):
        d = [self.csv.iloc[:, i - 1].to_list() for i in columns]
        return [list(i) for i in zip(*d)]  # 转置


if __name__ == '__main__':
    # reader = ExcelReader('example.xlsx')
    # print(reader.getSheetNames())
    # print(reader.getRows('P', 1))
    # print(reader.getColumnData('P', 1))
    # print(type(reader.getSheet('P')))
    # print(reader.getColumnData('Q1', 3))
    # s = CsvReader('../test0/雨量45+水情AB整合2.0.csv')
    # print(s.getColumns([1, -3, -2, -1, 0]))

    pass
