import os
import random
from typing import Sequence, Union

import matplotlib
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from matplotlib.colors import Normalize, LinearSegmentedColormap

matplotlib.use('TkAgg')
plt.rcParams['font.sans-serif'] = ['SimHei']
matplotlib.rcParams['axes.unicode_minus'] = False


def importMatplotlib():
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('TkAgg')  # 在不兼容的终端上使用python自带的窗口展示
    # plt.style.use('Solarize_Light2')  # 根据需要使用图像风格
    plt.rcParams.update({
        'axes.unicode_minus': False,  # 正确显示负号
        'font.serif': ['Times New Roman'],  # 优先使用新罗马Times New Roman字体
        'font.sans-serif': ['Microsoft YaHei'],  # 再使用宋体SimSun，等待官方解决该问题
        'mathtext.fontset': "custom",  # 数学字体
        'image.cmap': 'rainbow',  # 默认颜色方案为 rainbow
        'axes.titlesize': 12,  # 标题字体大小
        'legend.fontsize': 18,  # 图例字体大小
        'xtick.labelsize': 18,  # X轴刻度字体大小
        'ytick.labelsize': 18,  # Y轴刻度字体大小
        'axes.labelsize': 20,  # 轴标签字体大小
        # 'figure.figsize': (16, 9),   # 图形大小
        # 'lines.linewidth': 2,        # 线宽
        'lines.markersize': 8,  # 标记大小
        # 'grid.linestyle': '--',      # 网格线样式
        'font.size': 19,  # 默认字体大小，通常为text()方法中的字体大小
        'grid.color': 'gray',  # 网格颜色
        'grid.linewidth': 0.5,  # 网格线宽度
        # 'axes.facecolor': 'white',  # 轴背景色
        # 'figure.facecolor': 'white',  # 图形背景色
    })
    return plt


plt = importMatplotlib()
cs = [[0, "#452a3d", ], [0.2, "#44757a", ], [0.4, "#b7b5a0", ], [0.6, "#eed5b7", ], [0.8, "#d44c3c", ],
      [1, "#851321", ]]
cm = LinearSegmentedColormap.from_list('name', cs)


def readSheet(file: str, sheetName: str = None, onlyStr: bool = False) -> [[str]]:
    if not os.path.exists(file):
        raise FileNotFoundError(f'{file}文件不存在')
    try:
        workbook = openpyxl.load_workbook(file, read_only=False, data_only=True)
        sheetNames = list(workbook.sheetnames)
        if sheetName not in sheetNames:
            print(f"{file}文件中没有{sheetName}表格，将读取第一个表格: {sheetNames[0]}")
            sheetName = sheetNames[0]
        sheet = workbook[sheetName]
        data = []
        # 遍历工作表的每一行
        for row in sheet.iter_rows(values_only=True):
            row_data = []
            for cell_value in row:
                # 将单元格值转换为字符串，若为 None 则转换为空字符串
                if onlyStr:
                    cell_value = str(cell_value) if cell_value is not None else ""
                row_data.append(cell_value)
            data.append(row_data)
        workbook.close()
        print(f"读取{file}的{sheetName}成功")
        workbook.close()
        return data
    except Exception as e:
        raise Exception(f"读取{file}失败：{e}")


def drawSurface(ax: plt.Axes, xss: Sequence[Sequence[float]], yss: Sequence[Sequence[float]],
                zss: Sequence[Sequence[float]]):
    fig = ax.figure
    ax.remove()
    ax = fig.add_subplot(ax.get_subplotspec(), projection='3d')
    xss, yss, zss = np.array(xss), np.array(yss), np.array(zss)
    surf = ax.plot_surface(xss, yss, zss, cmap=cm, )
    bar = plt.colorbar(surf, shrink=0.5, aspect=8)
    norm = Normalize(vmin=zss.min(), vmax=zss.max())
    contour = ax.contour(xss, yss, zss, zdir='z', offset=np.min(zss), cmap=cm, norm=norm, zorder=0)
    ax.clabel(contour, inline=True, fontsize=8)
    ax.xaxis.set_major_locator(plt.MultipleLocator(5))
    ax.yaxis.set_major_locator(plt.MultipleLocator(5))
    ax.zaxis.set_major_locator(plt.MultipleLocator(5))
    # ax.set_title(f"$f(x_i)=x_i^4$")


def multipleLine(ax: plt.Axes, xs: Sequence, yss: Sequence[Sequence]):
    xs, yss = np.array(xs), np.array(yss)
    ls = ['EVO', 'KOA', 'ZOA', 'LSO', 'GRO', 'SWO', 'KOA']
    cs = cm(np.linspace(0, 1, len(ls)))
    for index, ys in enumerate(yss):
        ax.plot(xs, ys, linewidth=2, label=ls[index], color=cs[index])  # 线型
    ax.set_xlabel("迭代次数")
    ax.set_ylabel("资源耗费量")
    ax.grid()
    # ax.set_yticks([0, 0.2, 0.4, 0.6, 0.8, 1], ['$-9000$', '$-8000$', '$-7000$', '$-6000$', '$-5000$', '$-4000$'])
    ax.set_yticks([0, 0.25, 0.5, 0.75, 1], ['$10^{-15}$', '$10^{-10}$', '$10^{-5}$', '$10^{0}$', '$10^{5}$'])
    # ax.set_yticks([0, 0.25, 0.5, 0.75, 1], ['$1\\times 10^{-10}$', '$3\\times 10^{-10}$', '$5\\times 10^{-10}$', '$7\\times 10^{-10}$', '$9\\times 10^{-10}$'])
    ax.legend(loc='upper right')


def fxy(x: Union[float, np.ndarray], y: Union[float, np.ndarray]) -> Union[float, np.ndarray]:
    # z = x ** 2 - 10 * np.cos(2 * np.pi * x) + y ** 2 - 10 * np.cos(2 * np.pi * y) + 20
    # zf = lambda x, y: -x * np.sin(np.sqrt(np.abs(x))) - y * np.sin(np.sqrt(np.abs(y)))
    # z = np.array([zf(x_, y_) for x_, y_ in zip(x, y)])
    z = -20*np.exp(-0.2*np.sqrt(1/2*(x**2+y**2)))-np.exp(1/2*(np.cos(2*np.pi*x)+np.cos(2*np.pi*y)))+20+np.e
    return z


class ZRandom:
    @staticmethod
    def inverse(n, size=1000):
        def fx(n, y):
            result = n / (y + n)
            return result

        y = np.random.random(size) * 20
        return fx(n, y)


if __name__ == '__main__':
    file = r"C:\Users\刘高瞻\Desktop\工作簿1(2).xlsx"
    data = np.array(readSheet(file,'Sheet2')).T
    ran = np.random.random

    # d1 = ran(size=100)**0.1 * ran(size=100)**0.3
    # d2 = ran(size=100)**0.4 * ran(size=100)**1.5
    # d3 = ran(size=100)**0.2 * ran(size=100)**2.5
    # d4 = ran(size=100)**2 * ran(size=100)**0.5
    # d5 = ran(size=100)**0.8 * ran(size=100)**0.1
    # d6 = ran(size=100)**0.3 * ran(size=100)**0.4
    # data = [d1,d2,d3,d4,d5,d6]
    # for i in range(len(data)):
    #     d = data[i]
    #     d.sort()
    #     data[i] = d[::-1]
    # for i in data:
    #     for j in i:
    #         print(j,end=' ')
    #     print()

    xs = np.linspace(-10, 10, 300)
    ys = np.linspace(-10, 10, 300)
    X, Y = np.meshgrid(xs, ys)
    Z = fxy(X, Y)

    fig, axs = plt.subplots(1)
    drawSurface(axs, X, Y, Z)
    # multipleLine(axs, range(1, 101), data)
    plt.show()
