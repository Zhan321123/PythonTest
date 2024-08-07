# -*- coding: utf-8 -*-
"""
excel文件读取与写入
包含：
    .xlsx
    .csv

"""
import os
from typing import Sequence
import openpyxl
import pandas
import pandas as pd
from pandas import DataFrame


class ExcelReaderBase:
    """excel文件读取类，使用pandas库和openpyxl库辅助完成"""
    file: str
    workbook: dict[str, DataFrame]

    def __init__(self, file: str, header=None):
        if not file.endswith('.xlsx'):
            print('File is probably not an excel file')
        else:
            self.file = file
            self.workbook = pandas.read_excel(self.file, sheet_name=None, header=header)
            self._sheets = self.workbook.keys()

    def getSheetNames(self):
        """获取excel文件中的所有sheet名称"""
        pass

    def getSheetData(self, sheet: str):
        """获取一个表格中的二维列表内容"""
        pass


class ExcelReader(ExcelReaderBase):
    def __init__(self, file: str, header=None):
        super().__init__(file, header)

    def getSheetNames(self):
        return list(self.workbook.keys())

    def getSheetData(self, sheet: str):
        if sheet not in self._sheets:
            return None
        else:
            return self.workbook[sheet].to_numpy().tolist()

# TODO
class ExcelWriterBase:
    """excel文件写入类，使用openpyxl库辅助完成"""
    file: str
    workbook: openpyxl.Workbook

    def __init__(self, file: str):
        if file.endswith('.xlsx'):
            self.file = file
        else:
            self.file = file + '.xlsx'
        if os.path.exists(file):
            self.workbook = openpyxl.load_workbook(self.file, read_only=False, data_only=True)
        else:
            self.workbook = openpyxl.Workbook()
            self.workbook.save(self.file)

    def resetWorkbook(self, workbook: dict[str, DataFrame]):
        pass

    def appendSheet(self, sheet: str, data: DataFrame) -> bool:
        """
        append a sheet
        return dose the workbook has this sheet
        """
        pass


class ExcelWriter(ExcelWriterBase):

    def appendSheet(self, sheet: str, data: DataFrame) -> bool:
        if not self._sheetExists(sheet):
            df = pd.DataFrame(data)
            self.workbook.create_sheet(sheet)
            s = self.workbook.worksheets[sheet]
            for row in df.values:
                s.append(row.tolist())
            return True
        else:
            print('Sheet already exists')
            return False

    def modifyCellValue(self, sheet: str, row: int, column: int, value):
        """
        modify the workbook's sheet's row and colum's value
        """
        if not self._sheetExists(sheet):
            return
        self.workbook[sheet].cell(row=row, column=column).value = value

    def _sheetExists(self, sheet: str) -> bool:
        """
        this workbook has this sheet
        """
        return sheet in self.workbook.sheetnames

    def modifyRow(self, sheet: str, row: int, dataList: Sequence):
        """
        modify sheet's row's values
        """
        if not self._sheetExists(sheet):
            return
        for i in range(len(dataList)):
            self.workbook[sheet].cell(row=row, column=i + 1).value = dataList[i]

    def modifyColumn(self, sheet: str, column: int, dataList: Sequence):
        """
        modify sheet's column's values
        """
        if not self._sheetExists(sheet):
            return
        for i in range(len(dataList)):
            self.workbook[sheet].cell(row=i + 1, column=column).value = dataList[i]

    def save(self):
        """
        save workbook
        """
        self.workbook.save(self.file)

    def close(self):
        """
        close
        """
        self.workbook.close()


class CsvWriterBase:
    """csv文件写入类，使用pandas库辅助完成"""

    def __init__(self, data: list[list]):
        self.data = data

    def write(self, file: str = 'output.csv'):
        """写入文件"""
        pass


class CsvWriter(CsvWriterBase):
    def __init__(self, data: list[list]):
        super().__init__(data)

    def write(self, file: str = 'output.csv'):
        pandas.DataFrame(self.data).to_csv(file, header=False, index=False)


class CsvReaderBase:
    """csv文件读取类，使用pandas库辅助完成"""

    def __init__(self, file: str, header=None):
        self.file = file
        self.csv = pandas.read_csv(file, header=header, encoding='gbk', low_memory=False)
        self.rawData = self.csv.to_numpy().tolist()

    def getData(self) -> list[list]:
        """获取数据"""
        pass


class CsvReader(CsvReaderBase):
    def __init__(self, file: str, header=None):
        super().__init__(file, header)

    def getData(self) -> list[list]:
        return self.rawData


if __name__ == '__main__':
    pass
